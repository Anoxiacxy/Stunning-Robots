import os
from typing import Tuple, Dict

import numpy as np
from openpyxl import load_workbook, Workbook


def coordinate_to_position(coord: np.ndarray, playground_shape: np.ndarray) -> np.ndarray:
    """
    :param coord:
    :param playground_shape: like (n, m)
    :return:
    """
    x = coord[..., 0]
    y = coord[..., 1]
    # (..., n, 1)
    x_one_hot = np.eye(playground_shape[0], dtype=int)[x][..., np.newaxis]
    # (..., 1, m)
    y_one_hot = np.eye(playground_shape[1], dtype=int)[y][..., np.newaxis, :]
    # (..., n, m)
    return x_one_hot * y_one_hot


def position_to_coordinate(position: np.ndarray) -> np.ndarray:
    x = position.any(axis=-1).argmax(axis=-1)
    y = position.any(axis=-2).argmax(axis=-1)
    return np.stack((x, y), axis=-1)


def load_map_config(config_path):
    print(config_path)
    workbook = load_workbook(config_path)
    assert '机器人配置' in workbook.sheetnames
    assert '地图配置' in workbook.sheetnames
    assert '位置坐标' in workbook.sheetnames

    agents_sheet = workbook["机器人配置"]
    grids_sheet = workbook["地图配置"]
    coord_sheet = workbook["位置坐标"]

    base_cell = [3, 2]
    grids_size = [0, 0]
    while grids_sheet.cell(base_cell[0] + grids_size[0] + 1, base_cell[1]).value is not None:
        grids_size[0] += 1

    while grids_sheet.cell(base_cell[0], base_cell[1] + grids_size[1] + 1).value is not None:
        grids_size[1] += 1

    grids = [[grids_sheet.cell(base_cell[0] + 1 + i, base_cell[1] + 1 + j).value
              for j in range(grids_size[1])]
             for i in range(grids_size[0])]

    coord_map = dict()
    n_coord = 0
    while coord_sheet.cell(1 + n_coord + 1, 1).value is not None:
        n_coord += 1
        coord_map[coord_sheet.cell(n_coord + 1, 1).value] = [
            coord_sheet.cell(n_coord + 1, 2).value - 1,
            coord_sheet.cell(n_coord + 1, 3).value - 1
        ]

    n_agent = 0
    n_goal = 0
    while agents_sheet.cell(3 + n_goal + 1, 1).value is not None:
        n_goal += 1
    agents_speed = []
    agents_init_pos = []
    agents_dest_pos = []
    agents_periodicity = []
    speed_lcm = 1
    while agents_sheet.cell(1, 1 + n_agent + 1).value is not None:
        n_agent += 1
        speed = agents_sheet.cell(2, 1 + n_agent).value
        # print(speed)
        assert isinstance(speed, int) and speed >= 0
        speed_lcm = np.lcm(speed_lcm, speed) if speed > 0 else speed_lcm
        init_pos = coord_map[agents_sheet.cell(3, 1 + n_agent).value]
        dest_pos = []
        for i in range(n_goal):
            dest_pos.append(coord_map[agents_sheet.cell(4 + i, 1 + n_agent).value])
        agents_speed.append(speed)
        agents_dest_pos.append(dest_pos)
        agents_init_pos.append(init_pos)

    for i in range(n_agent):
        agents_periodicity.append(np.inf if agents_speed[i] == 0 else speed_lcm / agents_speed[i])

    return {
        "grids": grids,
        "n_agent": n_agent,
        "n_goal": n_goal,
        "fps": speed_lcm * 3,
        "periodicity": agents_periodicity,
        "goal_pos": agents_dest_pos,
        "init_pos": agents_init_pos,
    }


def save_rollout_xlsx(rollout_path, rollout_data: Dict):
    print(rollout_path)
    # 在内存创建一个工作簿
    workbook = Workbook()
    for page in rollout_data.keys():
        data = rollout_data[page]
        worksheet = workbook.create_sheet(page)
        for row in range(1, len(data) + 1):
            line = data[row - 1]
            for col in range(1, len(line) + 1):
                worksheet.cell(row=row, column=col).value = line[col - 1]
        workbook.active = worksheet
    # 工作簿保存到磁盘
    workbook.save(rollout_path)
